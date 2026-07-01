# TODO: add charts inside the excel sheets in next sprint
# FIXME: openpyxl column auto-width calculation fails for formula cells
# print("DEBUG: Excel writer initialized")
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Paths
ROOT = Path(__file__).resolve().parent.parent
PROCESSED = ROOT / 'data' / 'processed'
REPORTS = ROOT / 'reports'
REPORTS.mkdir(parents=True, exist_ok=True)

def main():
    print("Generating fund_scorecard_formatted.xlsx...")
    csv_path = PROCESSED / 'fund_scorecard.csv'
    if not csv_path.exists():
        raise FileNotFoundError("fund_scorecard.csv not found in processed data folder. Run analytics first.")
        
    df = pd.read_csv(csv_path)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Fund Scorecard"
    
    # Enable grid lines
    ws.views.sheetView[0].showGridLines = True
    
    # Colors (No Emojis, professional palette)
    HEADER_FILL = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    ALT_FILL = PatternFill(start_color="F7F9FA", end_color="F7F9FA", fill_type="solid")
    GREEN_FILL = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid") # soft green for top tier
    RED_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # soft red for bottom tier
    
    # Fonts
    HEADER_FONT = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    BODY_FONT = Font(name="Segoe UI", size=10)
    BOLD_FONT = Font(name="Segoe UI", size=10, bold=True)
    
    # Borders
    THIN_BORDER = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    # Write Headers
    headers = [
        "Rank", "AMFI Code", "Scheme Name", "Fund House", "Category", "Plan",
        "3Yr Return", "Sharpe Ratio", "Alpha", "Expense Ratio", "Max Drawdown", "Composite Score"
    ]
    ws.append(headers)
    
    # Format Headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        
    # Write Data
    for idx, r in df.iterrows():
        row_data = [
            int(r['score_rank']),
            int(r['amfi_code']),
            r['scheme_name'].split(' - ')[0],
            r['fund_house'],
            r['category'],
            r['plan'],
            float(r['return_3yr_pct']) / 100,
            float(r['sharpe_ratio']),
            float(r['alpha']) / 100 if 'alpha' in r else 0,
            float(r['expense_ratio_pct']) / 100,
            float(r['max_drawdown_pct']) / 100,
            float(r['composite_score'])
        ]
        ws.append(row_data)
        
    # Format Body Cells
    for row_idx in range(2, len(df) + 2):
        score_rank = ws.cell(row=row_idx, column=1).value
        comp_score = ws.cell(row=row_idx, column=12).value
        
        # Apply zebra striping
        row_fill = ALT_FILL if row_idx % 2 == 0 else PatternFill(fill_type=None)
        
        for col_idx in range(1, 13):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            
            # Alignments
            if col_idx in [1, 2]:
                cell.alignment = Alignment(horizontal="center")
            elif col_idx in [3, 4]:
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="right")
                
            # Number formats
            if col_idx in [7, 10, 11]:  # Percentages
                cell.number_format = '0.00%'
            elif col_idx in [8, 9, 12]:  # Floating numbers
                cell.number_format = '0.00'
                
            # Default Fill
            if row_fill.fill_type is not None:
                cell.fill = row_fill
                
            # Apply Conditional Fills based on scorecard values
            if col_idx == 12:  # Composite Score column
                cell.font = BOLD_FONT
                if comp_score >= 60:
                    cell.fill = GREEN_FILL
                elif comp_score < 40:
                    cell.fill = RED_FILL
            elif col_idx == 1:  # Rank column
                cell.font = BOLD_FONT
                if score_rank <= 5:
                    cell.fill = GREEN_FILL
                    
    # Adjust column widths dynamically
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            # Handle percentage string rendering length approximation
            val_str = str(cell.value or '')
            if cell.number_format == '0.00%':
                val_str = f"{val_str}%"
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
        
    # Save Workbook
    out_path = REPORTS / 'fund_scorecard_formatted.xlsx'
    wb.save(out_path)
    print(f"Saved fund_scorecard_formatted.xlsx")

if __name__ == '__main__':
    main()
